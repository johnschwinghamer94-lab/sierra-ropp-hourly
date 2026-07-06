#!/usr/bin/env python3
"""
cloud_hourly.py — runs in GitHub Actions (this PRIVATE repo) after Power Automate
pushes today's reports into hourly_reports/. Counts today's distinct ROPPs/TGLs and
publishes the SHARED hourly_state.json + hourly.json to the PUBLIC dashboard repo via
the API. Raw reports never leave this private repo. No machine needs to be on.
"""
import os, re, glob, json, base64
from datetime import datetime
import requests
try:
    from zoneinfo import ZoneInfo
    def _now(): return datetime.now(ZoneInfo("America/Los_Angeles"))
except Exception:
    def _now(): return datetime.utcnow()

TOKEN = os.environ["DASHBOARD_TOKEN"]
PUB = "johnschwinghamer94-lab/sierra-ropp-dashboard"
API = "https://api.github.com/repos/" + PUB + "/contents/"
H = {"Authorization": "token " + TOKEN, "Accept": "application/vnd.github+json"}


def _rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def _grouped(rows, jobcol):
    """One grouped row == one call/TGL; skips group headers and per-tech subtotal
    rows (job# < 6 digits). Mirrors UPDATE_DASHBOARD.iter_grouped exactly."""
    for r in rows[1:]:
        a = r[0] if r else None
        if isinstance(a, str) and a.strip().startswith("Assigned Technicians:"):
            continue
        if len(r) <= jobcol or r[jobcol] is None:
            continue
        jb = r[jobcol]
        s = str(int(jb)) if isinstance(jb, (int, float)) and float(jb).is_integer() else str(jb).strip()
        if not s.isdigit() or len(s) < 6:
            continue
        yield r


def _ts(p):
    # Power Automate prefixes each upload with a yyyyMMddHHmmss stamp, so the
    # leading digits sort chronologically. (git checkout gives every file the
    # same mtime, so mtime can't tell today's snapshots apart in the cloud.)
    m = re.match(r"(\d{14})", os.path.basename(p))
    return m.group(1) if m else "0"


def _find(*subs):
    cands = [p for p in glob.glob("hourly_reports/*.xlsx")
             if all(s in os.path.basename(p).lower() for s in subs)]
    return max(cands, key=lambda p: (_ts(p), os.path.getmtime(p))) if cands else None


def count_today():
    rev = _find("revenue")
    tgl = _find("tgls", "created") or _find("tgls")
    calls = sum(1 for _ in _grouped(_rows(rev), 3)) if rev else 0
    tgls  = sum(1 for _ in _grouped(_rows(tgl), 1)) if tgl else 0
    return calls, tgls


def rate(a, b):
    return round(a / b * 1000) / 10 if b else 0.0


def get(path):
    r = requests.get(API + path, headers=H)
    if r.status_code == 200:
        j = r.json()
        return json.loads(base64.b64decode(j["content"])), j["sha"]
    return None, None


def put(path, obj, sha, msg):
    body = {"message": msg, "content": base64.b64encode(json.dumps(obj).encode()).decode(), "branch": "main"}
    if sha:
        body["sha"] = sha
    r = requests.put(API + path, headers=H, json=body)
    r.raise_for_status()


def main():
    n = _now(); today = n.date().isoformat(); hh = f"{n.hour:02d}"
    calls, tgls = count_today()

    st, _ = get("hourly_state.json")               # shared state from public repo
    if not st or st.get("date") != today:
        st = {"date": today, "hours": {}}
    st["hours"][hh] = {"calls": calls, "tgls": tgls}

    hrs = sorted(st["hours"]); series = []; pc = pt = 0
    for h in hrs:
        c = st["hours"][h]["calls"]; t = st["hours"][h]["tgls"]
        series.append({"hour": h, "calls": c, "tgls": t, "rate": rate(t, c),
                       "dcalls": max(c-pc, 0), "dtgls": max(t-pt, 0), "drate": rate(max(t-pt, 0), max(c-pc, 0))})
        pc, pt = c, t
    latest = series[-1] if series else {"calls": 0, "tgls": 0, "rate": 0}
    out = {"date": today, "updated": n.strftime("%-I:%M %p"),
           "today": {"calls": latest["calls"], "tgls": latest["tgls"], "rate": latest["rate"]},
           "hours": series}

    _, ssha = get("hourly_state.json")
    put("hourly_state.json", st, ssha, f"Cloud hourly state {today} {hh}:00")
    _, osha = get("hourly.json")
    put("hourly.json", out, osha, f"Cloud hourly capture {today} {hh}:00")
    print(f"Cloud captured {today} {hh}:00 -> {calls} calls / {tgls} TGLs ({rate(tgls, calls)}%)")


if __name__ == "__main__":
    main()
