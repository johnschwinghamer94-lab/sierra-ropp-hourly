#!/usr/bin/env python3
"""
graph_hourly.py — FULLY-FREE cloud capture. Runs in GitHub Actions on a schedule
(no machine, no Power Automate premium). Reads today's ServiceTitan "today-only"
exports straight from OneDrive via Microsoft Graph (delegated auth), counts today's
distinct ROPPs/TGLs, and publishes the SHARED hourly_state.json + hourly.json to the
PUBLIC dashboard repo. Rotates its own Graph refresh-token secret so it never expires.

Secrets it needs (this private repo -> Settings -> Secrets -> Actions):
  GRAPH_CLIENT_ID     - Azure app (client) id
  GRAPH_TENANT_ID     - Azure directory (tenant) id
  GRAPH_REFRESH_TOKEN - from running graph_setup.py once (auto-rotated after that)
  DASHBOARD_TOKEN     - GitHub PAT with write to sierra-ropp-dashboard (+ this repo's secrets)
"""
import os, re, io, json, base64
from datetime import datetime
import requests
try:
    from zoneinfo import ZoneInfo
    def _now(): return datetime.now(ZoneInfo("America/Los_Angeles"))
except Exception:
    def _now(): return datetime.utcnow()

CLIENT = os.environ["GRAPH_CLIENT_ID"]
TENANT = os.environ["GRAPH_TENANT_ID"]
RTOKEN = os.environ["GRAPH_REFRESH_TOKEN"]
GHTOK  = os.environ["DASHBOARD_TOKEN"]
SELF   = os.environ.get("GITHUB_REPOSITORY", "johnschwinghamer94-lab/sierra-ropp-hourly")
PUB    = "johnschwinghamer94-lab/sierra-ropp-dashboard"
FOLDER = "CLAUDE STUFF/SILO_Reports"

GH = {"Authorization": "token " + GHTOK, "Accept": "application/vnd.github+json"}
PUBAPI = "https://api.github.com/repos/" + PUB + "/contents/"


# ---------- Microsoft Graph (delegated, refresh-token flow) ----------
def graph_token():
    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT}/oauth2/v2.0/token",
        data={"client_id": CLIENT, "grant_type": "refresh_token",
              "refresh_token": RTOKEN, "scope": "Files.Read offline_access"})
    r.raise_for_status()
    j = r.json()
    return j["access_token"], j.get("refresh_token")


def rotate_secret(new_rt):
    """Persist the freshly-issued refresh token back into this repo's secret so the
    login chain never lapses (Azure rolls the token on every use)."""
    if not new_rt or new_rt == RTOKEN:
        return
    try:
        from nacl import public, encoding
        pk = requests.get(f"https://api.github.com/repos/{SELF}/actions/secrets/public-key", headers=GH).json()
        sealed = public.SealedBox(public.PublicKey(pk["key"].encode(), encoding.Base64Encoder)).encrypt(new_rt.encode())
        requests.put(
            f"https://api.github.com/repos/{SELF}/actions/secrets/GRAPH_REFRESH_TOKEN",
            headers=GH,
            json={"encrypted_value": base64.b64encode(sealed).decode(), "key_id": pk["key_id"]}
        ).raise_for_status()
        print("Rotated GRAPH_REFRESH_TOKEN.")
    except Exception as e:
        print("WARN: could not rotate refresh token:", e)


def list_children(tok):
    from urllib.parse import quote
    url = (f"https://graph.microsoft.com/v1.0/me/drive/root:/{quote(FOLDER)}:/children"
           "?$top=400&$select=id,name,lastModifiedDateTime")
    items = []
    while url:
        j = requests.get(url, headers={"Authorization": "Bearer " + tok}).json()
        items += j.get("value", [])
        url = j.get("@odata.nextLink")
    return items


# ---------- pick today's snapshot & count ----------
def _range(name):
    m = re.search(r"(\d{2})_(\d{2})_(\d{2})\s*-\s*(\d{2})_(\d{2})_(\d{2})", name)
    if not m:
        return None, None
    try:
        from datetime import date
        s = date(2000+int(m.group(3)), int(m.group(1)), int(m.group(2)))
        e = date(2000+int(m.group(6)), int(m.group(4)), int(m.group(5)))
        return s, e
    except ValueError:
        return None, None


def today_file(items, *subs, exclude=()):
    """Newest 'today-only' (start==end==today) export whose name contains all subs
    and none of the `exclude` substrings."""
    today = _now().date()
    best = None
    for it in items:
        n = it["name"].lower()
        if not n.endswith(".xlsx") or not all(s in n for s in subs):
            continue
        if any(x in n for x in exclude):
            continue
        s, e = _range(it["name"])
        if s == today and e == today:
            if best is None or it["lastModifiedDateTime"] > best["lastModifiedDateTime"]:
                best = it
    return best


def rows(it, tok):
    from openpyxl import load_workbook
    data = requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/items/{it['id']}/content",
        headers={"Authorization": "Bearer " + tok}).content
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def _grouped(rows, jobcol):
    """Mirror UPDATE_DASHBOARD.iter_grouped: one row == one call/TGL. Skips the
    'Assigned Technicians:' group headers AND the per-tech subtotal / grand-total
    rows (those carry a small count like 1/2/7 in the job column). A real job#
    is a >=6-digit number, which is the exact filter the main dashboard uses."""
    for r in rows[1:]:
        a = r[0] if r else None
        if isinstance(a, str) and a.strip().startswith("Assigned Technicians:"):
            continue
        if len(r) <= jobcol:
            continue
        jb = r[jobcol]
        if jb is None:
            continue
        s = str(int(jb)) if isinstance(jb, (int, float)) and float(jb).is_integer() else str(jb).strip()
        if not s.isdigit() or len(s) < 6:
            continue
        yield r


def count(rev_rows, tgl_rows):
    calls = sum(1 for _ in _grouped(rev_rows, 3))   # Revenue: Job# in col 3
    tgls  = sum(1 for _ in _grouped(tgl_rows, 1))    # TGLs Created: Job# in col 1
    return calls, tgls


def _fnum(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def per_tech(rev_rows, tgl_rows):
    """Today's calls/TGLs/revenue per technician (tech name: Revenue col 7, TGLs
    col 3; TGL revenue = TGLs 'Sales from Leads Created' col 8; first-listed on
    shared jobs). Sorted by TGLs desc, then calls desc."""
    def tally(rows, jobcol, techcol, sumcol=None):
        cnt = {}; amt = {}
        for r in _grouped(rows, jobcol):
            t = r[techcol] if len(r) > techcol else None
            name = str(t).split(",")[0].strip() if t else "Unassigned"
            cnt[name] = cnt.get(name, 0) + 1
            if sumcol is not None:
                amt[name] = amt.get(name, 0.0) + (_fnum(r[sumcol]) if len(r) > sumcol else 0.0)
        return cnt, amt
    rc, _ = tally(rev_rows, 3, 7)
    tc, rv = tally(tgl_rows, 1, 3, 8)
    techs = [{"name": n, "calls": rc.get(n, 0), "tgls": tc.get(n, 0),
              "rate": rate(tc.get(n, 0), rc.get(n, 0)), "rev": round(rv.get(n, 0.0))}
             for n in (set(rc) | set(tc))]
    techs.sort(key=lambda x: (-x["tgls"], -x["calls"], x["name"]))
    return techs


def _asdate(v):
    from datetime import datetime as _dt, date as _date
    if isinstance(v, _dt): return v.date()
    if isinstance(v, _date): return v
    try: return _dt.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError): return None


def sched_metrics(rows):
    """Same-day / ran / sold from the today-only 'Scheduled vs Ran vs Sold' report.
    Grouped by sales tech (col 3) but ATTRIBUTED to the ROPP source tech (col 4), like
    the dashboard. col6 Scheduled Date, col7 Created Date (same-day flip = equal),
    col10 'Jobs Estimate Sales Subtotal' = SOLD $ (sold when > 0). NOT col8 'Installed',
    which is $0 until physically installed -- Subtotal>0 == ServiceTitan Closed=True."""
    per = {}; ran = same = sold = 0; soldrev = 0.0
    for r in _grouped(rows, 1):
        name = str(r[4]).split(",")[0].strip() if len(r) > 4 and r[4] else "Unassigned"
        sd = _asdate(r[6]) if len(r) > 6 else None
        cr = _asdate(r[7]) if len(r) > 7 else None
        sold_amt = _fnum(r[10]) if len(r) > 10 else 0.0   # col 10 = Estimate Sales Subtotal = SOLD price
        e = per.setdefault(name, {"name": name, "ran": 0, "same": 0, "sold": 0, "soldrev": 0.0})
        e["ran"] += 1; ran += 1
        if sd and cr and sd == cr: e["same"] += 1; same += 1
        if sold_amt > 0: e["sold"] += 1; e["soldrev"] += sold_amt; sold += 1; soldrev += sold_amt
    techs = sorted(per.values(), key=lambda x: (-x["ran"], -x["same"], x["name"]))
    for t in techs: t["soldrev"] = round(t["soldrev"]); t["samePct"] = rate(t["same"], t["ran"])
    return {"ran": ran, "same": same, "samePct": rate(same, ran),
            "sold": sold, "soldrev": round(soldrev), "closeRate": rate(sold, ran), "techs": techs}


def rate(a, b):
    return round(a / b * 1000) / 10 if b else 0.0


# ---------- publish shared state to the public repo ----------
def pget(path):
    r = requests.get(PUBAPI + path, headers=GH)
    if r.status_code == 200:
        j = r.json()
        return json.loads(base64.b64decode(j["content"])), j["sha"]
    return None, None


def pput(path, obj, sha, msg):
    body = {"message": msg, "content": base64.b64encode(json.dumps(obj).encode()).decode(), "branch": "main"}
    if sha:
        body["sha"] = sha
    requests.put(PUBAPI + path, headers=GH, json=body).raise_for_status()


def _ensure_st_creds():
    """Materialize ~/.servicetitan/sierra.json from the ST_CREDS_JSON env (GitHub secret)
    if it isn't already on disk. Returns True if usable creds are present."""
    import pathlib
    fp = pathlib.Path.home() / ".servicetitan" / "sierra.json"
    if fp.exists():
        return True
    raw = os.environ.get("ST_CREDS_JSON", "")
    if not raw.strip():
        return False
    try:
        json.loads(raw)
    except Exception:
        return False
    fp.parent.mkdir(parents=True, exist_ok=True); fp.write_text(raw); os.chmod(fp, 0o600)
    return True


def _st_today(today):
    """Today's Revenue / TGLs-Created / Scheduled rows straight from the ServiceTitan
    Reporting API (same reports the today-only Excel exports come from), in the Excel
    row shape the counters expect. Fresher than waiting on the OneDrive export to land."""
    import ropp_live as RL
    return (RL.fetch_report_rows("Revenue_By_JobType.xlsx", today, today),
            RL.fetch_report_rows("ROPP_TGLs_Created.xlsx",   today, today),
            RL.fetch_report_rows("ROPP_TGLs_Scheduled.xlsx", today, today))


def _entity_tgls_today(today):
    """John's rule (2026-07-16): a TGL counts on the day its TICKET IS CREATED —
    including follow-up turnovers on calls that ran earlier. The TGLs-Created
    report can't express that via the API (LeadCreated comes back empty and
    DateType filters on the call's ScheduledDate), so count lead JOBS from the
    entity API: created today, job type in the "Estimate … TGL" family, with the
    dup-ticket rule (canceled leads on a multi-lead call are duplicates; a lone
    canceled lead is a real canceled TGL and still counts as created).
    Returns (total, {tech_full_name: count})."""
    import st_client as st
    from datetime import datetime as _dt, time as _time, timezone as _tz

    day0 = _dt.combine(_dt.strptime(today, "%Y-%m-%d").date(), _time.min
                       ).astimezone(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def paged(path, params):
        page = 1
        while True:
            r = st.api_get(path, {**params, "page": page, "pageSize": 200})
            yield from r.get("data", [])
            if not r.get("hasMore"):
                break
            page += 1

    jts = {x["id"]: x.get("name", "") for x in paged("/jpm/v2/tenant/{tenant}/job-types", {})}
    leads = []
    for lj in paged("/jpm/v2/tenant/{tenant}/jobs", {"createdOnOrAfter": day0}):
        gls = lj.get("jobGeneratedLeadSource") or {}
        jtn = jts.get(lj.get("jobTypeId")) or ""
        if gls.get("jobId") and jtn.startswith("Estimate") and "TGL" in jtn:
            leads.append({"src": gls["jobId"], "emp": gls.get("employeeId"),
                          "can": lj.get("jobStatus") == "Canceled"})
    by_src = {}
    for L in leads:
        by_src.setdefault(L["src"], []).append(L)
    leads = []
    for grp in by_src.values():
        live = [L for L in grp if not L["can"]]
        leads.extend(live if live else grp[:1])

    eids = list({L["emp"] for L in leads if L["emp"]})
    emps = {}
    for i in range(0, len(eids), 50):
        for e in paged("/settings/v2/tenant/{tenant}/employees",
                       {"ids": ",".join(map(str, eids[i:i+50]))}):
            emps[e["id"]] = e.get("name", "")
    by_tech = {}
    for L in leads:
        name = emps.get(L["emp"]) or None
        if not name:
            try:
                r = st.api_get("/dispatch/v2/tenant/{tenant}/appointment-assignments",
                               {"jobId": L["src"], "pageSize": 5})
                names = [a.get("technicianName") for a in r.get("data", [])
                         if a.get("active") and a.get("technicianName")]
                name = names[0] if names else "Unassigned"
            except Exception:
                name = "Unassigned"
        by_tech[name] = by_tech.get(name, 0) + 1
    return len(leads), by_tech


def main():
    n = _now(); today = n.date().isoformat(); hh = f"{n.hour:02d}"
    rev_rows = tgl_rows = sch_rows = None

    # PREFERRED: pull today's reports live from the ServiceTitan API.
    if _ensure_st_creds():
        try:
            rev_rows, tgl_rows, sch_rows = _st_today(today)
            print("Today source: ServiceTitan Reporting API")
        except Exception as e:
            print(f"ServiceTitan today-fetch failed ({e}); falling back to OneDrive Excel")
            rev_rows = None

    # FALLBACK: today-only Excel exports from OneDrive via Graph (original behavior).
    if rev_rows is None:
        tok, new_rt = graph_token()
        rotate_secret(new_rt)
        items = list_children(tok)
        rev = today_file(items, "revenue")
        tgl = today_file(items, "tgls", "created") or today_file(items, "tgls", exclude=("scheduled", "ran", "sold", "estimate"))
        if not rev and not tgl:
            print("No today-only revenue/tgls export in OneDrive yet; nothing to publish.")
            return
        rev_rows = rows(rev, tok) if rev else []
        tgl_rows = rows(tgl, tok) if tgl else []
        sch = today_file(items, "scheduled")
        sch_rows = rows(sch, tok) if sch else None
        print("Today source: OneDrive Excel (today-only exports)")

    calls, tgls = count(rev_rows, tgl_rows)
    techs = per_tech(rev_rows, tgl_rows)
    # TGLs count on their CREATION day (John's rule) — override the report's
    # scheduled-date lens with the entity-API count; report stays the fallback.
    try:
        e_tot, e_by = _entity_tgls_today(today)
        tgls = e_tot
        seen = set()
        for tec in techs:
            tec["tgls"] = e_by.get(tec["name"], 0)
            tec["rate"] = rate(tec["tgls"], tec["calls"])
            seen.add(tec["name"])
        for enm, ecnt in e_by.items():
            if enm not in seen:
                techs.append({"name": enm, "calls": 0, "tgls": ecnt, "rate": None, "rev": 0})
        techs = [tec for tec in techs if tec["calls"] or tec["tgls"]]
        techs.sort(key=lambda x: (-x["tgls"], -x["calls"], x["name"]))
        print(f"TGL lens: entity created-today ({e_tot})")
    except Exception as e:
        print(f"entity TGL count failed ({e}); keeping report lens")
    sched = sched_metrics(sch_rows) if sch_rows else None

    st, _ = pget("hourly_state.json")
    if not st or st.get("date") != today:
        st = {"date": today, "hours": {}}
    st["hours"][hh] = {"calls": calls, "tgls": tgls}

    hrs = sorted(st["hours"]); series = []; pc = pt = 0
    for h in hrs:
        c = st["hours"][h]["calls"]; t = st["hours"][h]["tgls"]
        series.append({"hour": h, "calls": c, "tgls": t, "rate": rate(t, c),
                       "dcalls": max(c-pc, 0), "dtgls": max(t-pt, 0), "drate": rate(max(t-pt, 0), max(c-pc, 0))})
        pc, pt = c, t
    latest = series[-1] if series else {"calls": 0, "tgls": 0, "rate": 0}
    revenue = sum(t["rev"] for t in techs)
    today_block = {"calls": latest["calls"], "tgls": latest["tgls"], "rate": latest["rate"],
                   "rev": revenue, "techs": techs}
    if sched:
        today_block["sched"] = sched
    out = {"date": today, "updated": n.strftime("%I:%M %p").lstrip("0"),
           "generatedMs": int(n.timestamp() * 1000),   # health banner reads this
           "today": today_block, "hours": series}

    _, ssha = pget("hourly_state.json")
    pput("hourly_state.json", st, ssha, f"Cloud(graph) hourly state {today} {hh}:00")
    _, osha = pget("hourly.json")
    pput("hourly.json", out, osha, f"Cloud(graph) hourly capture {today} {hh}:00")
    print(f"Published {today} {hh}:00 -> {calls} calls / {tgls} TGLs ({rate(tgls, calls)}%)")

    # Kick the full ROPP rebuild (daily.yml) every quarter hour so MTD/YTD stay near-live.
    # GitHub's native cron skips ticks on low-traffic repos, but this cron-job.org-driven
    # job runs dependably; the first tick inside each quarter (:00/:15/:30/:45) dispatches
    # the heavier full rebuild. Idempotent — auto_update_dashboard skips its push when
    # nothing changed, and daily.yml's own cron stays as an overnight backup.
    if _now().minute % 15 < 3:
        try:
            r = requests.post(
                f"https://api.github.com/repos/{SELF}/actions/workflows/daily.yml/dispatches",
                headers=GH, json={"ref": "main"})
            print(f"Kicked daily.yml full rebuild -> HTTP {r.status_code}")
        except Exception as e:
            print("WARN: could not dispatch daily.yml:", e)


if __name__ == "__main__":
    main()
