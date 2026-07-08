"""
UPDATE_DASHBOARD.py  -  ROPP DEPARTMENT Dashboard rebuilder
===========================================================
Department-wide everywhere; SILO ONLY tab stays the SILO roster.
NO tech filter — every technician that appears on the reports is included.

Rules:
  * Shared tickets: a SILO member on the ticket keeps it; otherwise the
    first-listed tech gets it (one ticket -> one tech).
  * Revenue from TGL Created "Sales from Leads Created".
  * Service vs Maintenance by Business Unit text.
  * Cancellations: YTD by cancelled date, month/week by scheduled date.
  * Same-day/next-day keyed to Lead-Generated-From-Source-Technician.
  * Andrew Alonso pre-2026-06-14 excluded from the TEAM A/B split ONLY.
"""

import os, re, json, calendar, argparse
from datetime import datetime, date, timedelta
from pathlib import Path
from collections import defaultdict
import warnings; warnings.filterwarnings("ignore")

SCRIPT_DIR  = Path(__file__).parent
REPORTS_DIR = SCRIPT_DIR / "SILO_Reports"

# ---- SILO roster (curated) --------------------------------------------------
TEAM_A = ["Noah Weng", "Joe Mendoza", "Benjamin Wyllie", "Nikko April",
          "Andrew Trujillo", "Dustin Romine", "Juan Tlatenchi"]
TEAM_B = ["Brandon Moreno", "Francisco Valencia", "Mario Castro", "Cole Pantol",
          "Nathan Colquitt", "Andrew Alonso", "Robert Silinzy"]
ALEX   = "Alex - Oleksiy Yakovchuk"
SILO   = [ALEX] + TEAM_A + TEAM_B           # 15 (SILO_SET, incl Andrew Alonso)
SILO_12 = [n for n in SILO if n != "Andrew Alonso"]   # 14 (SILO_12 const)

ROSTER_START = {"Andrew Alonso": date(2026, 6, 14)}    # team-split tab only

YEAR        = 2026
YEAR_START  = date(YEAR, 1, 1)
YEAR_END    = date(YEAR, 12, 31)
DAYS_TOTAL  = (YEAR_END - YEAR_START).days + 1
TECH_GOAL   = 3_500_000
DEPT_GOAL   = 30_000_000
MONTH_NAMES = ["January","February","March","April","May","June","July",
               "August","September","October","November","December"]

# The Power Automate flow saves each ServiceTitan email attachment under its
# native (dated) name, e.g.
#   "ROPP TGLS CREATED (JOHN)_Dated 01_01_26 - 06_30_26.xlsx"
#   "ROPP CANCELATIONS (JOHN)_Dated 01_01_26 - 06_30_26.xlsx"
#   "TGLS SCHEDULED DATE AND TGLS RAN DATE VS SOLD _Dated 01_01_26 - 06_30_26.xlsx"
#   "Johns Copy of Ericka's Revenue by Job Type_Dated 01_01_26 - 06_30_26.xlsx"
# Older tooling wrote the plain canonical name (e.g. "ROPP_TGLs_Created.xlsx").
# We map each canonical report to a distinctive phrase and pick the newest file
# whose (separator-normalised) name contains that phrase -- this matches BOTH the
# new dated names and the legacy canonical names, so nothing breaks in between.
REPORT_MATCH = {
    "ROPP_TGLs_Created.xlsx":   "tgls created",
    "ROPP_TGLs_Scheduled.xlsx": "tgls scheduled",
    "ROPP_Cancelations.xlsx":   "cancelations",
    "Revenue_By_JobType.xlsx":  "revenue by job",
}

def _norm(s):
    # lower-case and collapse runs of spaces/underscores to a single space so
    # "ROPP_TGLs_Created" and "ROPP TGLS CREATED (JOHN)" both normalise to a
    # form containing "tgls created".
    return re.sub(r"[\s_]+", " ", s.lower())

def _range_start(name):
    """Parse the 'Dated MM_DD_YY - MM_DD_YY' span start from a report filename.
    Returns the start date, or None if not parseable."""
    m = re.search(r"(\d{2})_(\d{2})_(\d{2})\s*-\s*(\d{2})_(\d{2})_(\d{2})", name)
    if not m:
        return None
    try:
        return date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None

def resolve_report(fname):
    phrase = REPORT_MATCH.get(fname)
    cands = []
    for p in REPORTS_DIR.glob("*.xlsx"):
        if p.name.startswith("~$"):        # skip Excel lock/temp files
            continue
        if phrase and phrase in _norm(p.name):
            cands.append(p)
        elif p.name == fname:
            cands.append(p)
    if cands:
        # The daily pipeline needs the FULL-YEAR (YTD) report. Hourly "Today"-filtered
        # exports (start==today) must NOT win here or YTD collapses to one day.
        # Prefer the earliest span start (widest range); tiebreak newest mtime.
        return min(cands, key=lambda p: (_range_start(p.name) or date(9999, 1, 1), -p.stat().st_mtime))
    exact = REPORTS_DIR / fname
    if exact.exists():
        return exact
    raise FileNotFoundError("No file found for report '%s' (match phrase %r) in %s"
                            % (fname, phrase, REPORTS_DIR))

def load_rows(fname):
    from openpyxl import load_workbook
    wb = load_workbook(resolve_report(fname), data_only=True, read_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]

def names_of(name):
    if not name: return []
    return [p.strip() for p in str(name).split(",") if p.strip()]

def resolve(name):
    parts = names_of(name)
    for p in parts:
        if p in SILO: return p
    return parts[0] if parts else None

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    try:    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except: return None

def fnum(v):
    try:    return float(v)
    except: return 0.0

def is_maint(bu): return "Maintenance" in str(bu or "")
def is_svc(bu):   return "Service" in str(bu or "")

def iter_grouped(rows, prefix, jobcol):
    cur = None
    for r in rows[1:]:
        a = r[0]
        if isinstance(a, str) and a.strip().startswith(prefix + ":"):
            cur = a.split(":", 1)[1].strip(); continue
        if len(r) <= jobcol: continue
        jb = r[jobcol]
        if jb is None: continue
        s = str(jb).strip()
        if not s.isdigit() or len(s) < 6: continue
        yield cur, r

def week_of_month(d): return "W" + str(min(4, (d.day - 1) // 7 + 1))
def conv_week_labels(mo):
    s = MONTH_NAMES[mo-1][:3]
    return [s+" 1-7", s+" 8-14", s+" 15-21", s+" 22-28"]
def month_week_labels(yr, mo):
    last = calendar.monthrange(yr, mo)[1]; s = MONTH_NAMES[mo-1][:3]
    return {"W1": s+" 1-7","W2": s+" 8-14","W3": s+" 15-21","W4": s+" 22-"+str(last)}

def rate(g, c): return round(g / c * 100, 1) if c else 0.0

# =============================================================================
#  PARSE  (all techs)
# =============================================================================
def blank():
    return {"ytd": dict(calls=0,tgls=0,revenue=0.0,svc_calls=0,svc_tgls=0,maint_calls=0,maint_tgls=0),
            "mtd": dict(calls=0,tgls=0,revenue=0.0,svc_calls=0,svc_tgls=0,maint_calls=0,maint_tgls=0),
            "monthly": {m: dict(calls=0,tgls=0,revenue=0.0) for m in MONTH_NAMES}}

def jobkey(v):
    if v is None: return None
    return str(int(v)) if isinstance(v, (int, float)) and float(v).is_integer() else str(v).strip()

def sub_by_srcjob():
    """TGL revenue = sold Estimate Sales Subtotal (col 10) from the Scheduled-vs-Ran-vs-Sold
    report, summed per lead-source job number (col 5), joined to ROPP_TGLs_Created Job#."""
    d = {}
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Scheduled.xlsx"), "Assigned Technicians", 1):
        k = jobkey(r[5])
        if k: d[k] = d.get(k, 0.0) + (fnum(r[10]) if len(r) > 10 else 0.0)
    return d

def parse_all(today):
    techs = defaultdict(blank)
    SUB = sub_by_srcjob()
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Created.xlsx"), "Assigned Technicians", 1):
        tech = resolve(r[3]) or resolve(grp); d = to_date(r[5])
        if not tech or d is None or d.year != YEAR: continue
        rev = SUB.get(jobkey(r[1]), 0.0); bu = r[4]; t = techs[tech]; mo = MONTH_NAMES[d.month-1]
        t["ytd"]["tgls"] += 1; t["ytd"]["revenue"] += rev
        t["monthly"][mo]["tgls"] += 1; t["monthly"][mo]["revenue"] += rev
        if is_maint(bu): t["ytd"]["maint_tgls"] += 1
        elif is_svc(bu): t["ytd"]["svc_tgls"] += 1
        if d.month == today.month:
            t["mtd"]["tgls"] += 1; t["mtd"]["revenue"] += rev
            if is_maint(bu): t["mtd"]["maint_tgls"] += 1
            elif is_svc(bu): t["mtd"]["svc_tgls"] += 1
    for grp, r in iter_grouped(load_rows("Revenue_By_JobType.xlsx"), "Assigned Technicians", 3):
        tech = resolve(r[7]) or resolve(grp); d = to_date(r[4])
        if not tech or d is None or d.year != YEAR: continue
        bu = r[10]; t = techs[tech]; mo = MONTH_NAMES[d.month-1]
        t["ytd"]["calls"] += 1; t["monthly"][mo]["calls"] += 1
        if is_maint(bu): t["ytd"]["maint_calls"] += 1
        elif is_svc(bu): t["ytd"]["svc_calls"] += 1
        if d.month == today.month:
            t["mtd"]["calls"] += 1
            if is_maint(bu): t["mtd"]["maint_calls"] += 1
            elif is_svc(bu): t["mtd"]["svc_calls"] += 1
    for t in techs.values():
        for p in ("ytd","mtd"):
            t[p]["revenue"] = round(t[p]["revenue"])
            t[p]["rate"] = rate(t[p]["tgls"], t[p]["calls"])
        for m in MONTH_NAMES: t["monthly"][m]["revenue"] = round(t["monthly"][m]["revenue"])

    cancel = defaultdict(lambda: {"ytd":0,"monthly":defaultdict(int),"weekly":defaultdict(int)})
    for grp, r in iter_grouped(load_rows("ROPP_Cancelations.xlsx"), "Lead Generated By", 1):
        tech = resolve(r[10]) or resolve(grp)
        cd = to_date(r[9]); sdt = to_date(r[7])
        if not tech: continue
        if cd is not None and cd.year == YEAR: cancel[tech]["ytd"] += 1
        if sdt is not None and sdt.year == YEAR:
            cancel[tech]["monthly"][MONTH_NAMES[sdt.month-1]] += 1
            if sdt.month == today.month: cancel[tech]["weekly"][week_of_month(sdt)] += 1

    sd = defaultdict(lambda: {"ytd":{"total":0,"flipped":0,"nextday":0},
                              "monthly":defaultdict(lambda:{"total":0,"flipped":0,"nextday":0}),
                              "weekly":defaultdict(lambda:{"total":0,"flipped":0,"nextday":0})})
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Scheduled.xlsx"), "Assigned Technicians", 1):
        tech = resolve(r[4]); ran = to_date(r[6]); cr = to_date(r[7])
        if not tech or ran is None or ran.year != YEAR: continue
        same = cr is not None and ran == cr; nxt = cr is not None and (ran-cr).days == 1
        mo = MONTH_NAMES[ran.month-1]
        for b in (sd[tech]["ytd"], sd[tech]["monthly"][mo]):
            b["total"] += 1; b["flipped"] += 1 if same else 0; b["nextday"] += 1 if nxt else 0
        if ran.month == today.month:
            b = sd[tech]["weekly"][week_of_month(ran)]
            b["total"] += 1; b["flipped"] += 1 if same else 0; b["nextday"] += 1 if nxt else 0
    return techs, cancel, sd

# =============================================================================
#  BUILD BLOCKS
# =============================================================================
def build_initial(techs):
    return {n: {"ytd": dict(t["ytd"]), "mtd": dict(t["mtd"]),
                "monthly": {m: dict(calls=t["monthly"][m]["calls"], tgls=t["monthly"][m]["tgls"])
                            for m in MONTH_NAMES}}
            for n, t in techs.items() if "," not in n}

def build_cancel(techs, cancel, today, months):
    ytd, monthly, weekly = {}, {}, {}
    for n in set(list(techs.keys()) + list(cancel.keys())):
        if "," in n: continue
        t = techs.get(n)
        ytd[n] = {"scheduled": t["ytd"]["tgls"] if t else 0, "cancelled": cancel[n]["ytd"]}
        monthly[n] = {m: {"scheduled": t["monthly"][m]["tgls"] if t else 0,
                          "cancelled": cancel[n]["monthly"].get(m,0)} for m in months}
        weekly[n] = {w: {"scheduled":0,"cancelled":cancel[n]["weekly"].get(w,0)} for w in ("W1","W2","W3","W4")}
    return {"ytd":ytd,"monthly":monthly,"weekly":weekly,"months":months,
            "weeks":["W1","W2","W3","W4"],"weekLabels":month_week_labels(YEAR,today.month)}

def build_sameday(sd, today, months, all_names):
    ytd, monthly, weekly = {}, {}, {}
    for n, s in sd.items():
        if "," in n: continue
        ytd[n] = dict(s["ytd"])
        monthly[n] = {m: dict(s["monthly"].get(m,{"total":0,"flipped":0,"nextday":0})) for m in months}
        weekly[n] = {w: dict(s["weekly"].get(w,{"total":0,"flipped":0,"nextday":0})) for w in ("W1","W2","W3","W4")}
    for n in all_names:
        if "," in n or n in ytd: continue
        ytd[n] = {"total":0,"flipped":0,"nextday":0}
        monthly[n] = {m: {"total":0,"flipped":0,"nextday":0} for m in months}
        weekly[n] = {w: {"total":0,"flipped":0,"nextday":0} for w in ("W1","W2","W3","W4")}
    return {"ytd":ytd,"monthly":monthly,"weekly":weekly,"months":months,
            "weeks":["W1","W2","W3","W4"],"weekLabels":month_week_labels(YEAR,today.month)}

def build_weekly_prevmonth(today):
    pm = today.month - 1 or 12; py = YEAR if today.month > 1 else YEAR-1
    agg = defaultdict(lambda: {w: dict(calls=0,tgls=0,revenue=0.0) for w in ("W1","W2","W3","W4")})
    SUB = sub_by_srcjob()
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Created.xlsx"), "Assigned Technicians", 1):
        tech = resolve(r[3]) or resolve(grp); d = to_date(r[5])
        if not tech or d is None or d.year != py or d.month != pm: continue
        w = week_of_month(d); agg[tech][w]["tgls"] += 1; agg[tech][w]["revenue"] += SUB.get(jobkey(r[1]), 0.0)
    for grp, r in iter_grouped(load_rows("Revenue_By_JobType.xlsx"), "Assigned Technicians", 3):
        tech = resolve(r[7]) or resolve(grp); d = to_date(r[4])
        if not tech or d is None or d.year != py or d.month != pm: continue
        agg[tech][week_of_month(d)]["calls"] += 1
    out = {}
    for n, wk in agg.items():
        if "," in n: continue
        out[n] = {w: dict(calls=wk[w]["calls"],tgls=wk[w]["tgls"],revenue=round(wk[w]["revenue"]),
                          rate=rate(wk[w]["tgls"],wk[w]["calls"])) for w in ("W1","W2","W3","W4")}
    return {"weeks":["W1","W2","W3","W4"],"weekLabels":month_week_labels(py,pm),"techs":out}

def build_weekly_conv(today, all_names):
    labels = conv_week_labels(today.month)
    idx = {"W1":0,"W2":1,"W3":2,"W4":3}
    per = defaultdict(lambda: [dict(calls=0,tgls=0,revenue=0.0) for _ in range(4)])
    SUB = sub_by_srcjob()
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Created.xlsx"), "Assigned Technicians", 1):
        tech = resolve(r[3]) or resolve(grp); d = to_date(r[5])
        if not tech or d is None or d.year != YEAR or d.month != today.month: continue
        i = idx[week_of_month(d)]; per[tech][i]["tgls"] += 1; per[tech][i]["revenue"] += SUB.get(jobkey(r[1]), 0.0)
    for grp, r in iter_grouped(load_rows("Revenue_By_JobType.xlsx"), "Assigned Technicians", 3):
        tech = resolve(r[7]) or resolve(grp); d = to_date(r[4])
        if not tech or d is None or d.year != YEAR or d.month != today.month: continue
        per[tech][idx[week_of_month(d)]]["calls"] += 1
    techs_out = {}; tot = [dict(calls=0,tgls=0,revenue=0.0) for _ in range(4)]
    for n, weeks in per.items():
        if "," in n: continue
        techs_out[n] = {labels[i]: dict(calls=weeks[i]["calls"],tgls=weeks[i]["tgls"],
                        revenue=round(weeks[i]["revenue"]),rate=rate(weeks[i]["tgls"],weeks[i]["calls"])) for i in range(4)}
        for i in range(4):
            tot[i]["calls"]+=weeks[i]["calls"]; tot[i]["tgls"]+=weeks[i]["tgls"]; tot[i]["revenue"]+=weeks[i]["revenue"]
    for n in all_names:
        if "," in n or n in techs_out: continue
        techs_out[n] = {labels[i]: dict(calls=0,tgls=0,revenue=0,rate=0.0) for i in range(4)}
    team_totals = [dict(week=labels[i],calls=tot[i]["calls"],tgls=tot[i]["tgls"],
                        rate=rate(tot[i]["tgls"],tot[i]["calls"]),revenue=round(tot[i]["revenue"])) for i in range(4)]
    return {"month":MONTH_NAMES[today.month-1],"weeks":labels,"team_totals":team_totals,"techs":techs_out}

def build_pace(techs, today, de):
    dr = DAYS_TOTAL - de; out = {}
    for n, t in techs.items():
        if "," in n: continue
        rv = t["ytd"]["revenue"]; exp = round(TECH_GOAL*de/DAYS_TOTAL); da = round(rv/de) if de else 0
        out[n] = {"goal":TECH_GOAL,"ytd_revenue":rv,"days_elapsed":de,"days_total":DAYS_TOTAL,
                  "days_remaining":dr,"expected_pace":exp,"daily_needed":round((TECH_GOAL-rv)/dr) if dr else 0,
                  "daily_actual":da,"projected_eoy":round(rv+da*dr),"pace_pct":round(rv/exp*100,1) if exp else 0,
                  "monthly":[dict(month=m,revenue=t["monthly"][m]["revenue"],tgls=t["monthly"][m]["tgls"],
                                  calls=t["monthly"][m]["calls"]) for m in MONTH_NAMES[:today.month]]}
    return out

def build_monthly_detail(techs, months):
    out = {}
    for m in months:
        out[m] = {}
        for n, t in techs.items():
            if "," in n: continue
            mm = t["monthly"][m]
            if mm["calls"]==0 and mm["tgls"]==0 and mm["revenue"]==0: continue
            out[m][n] = dict(calls=mm["calls"],svc_calls=0,maint_calls=0,svc_tgls=0,maint_tgls=0,
                             total_tgls=mm["tgls"],svc_rev=0,maint_rev=0,total_rev=mm["revenue"])
    return out

def build_allteams(techs):
    out = []
    for n, t in techs.items():
        if "," in n: continue
        out.append(dict(name=n,is_silo=(n in SILO),ytd_calls=t["ytd"]["calls"],ytd_tgls=t["ytd"]["tgls"],
                        ytd_rate=t["ytd"]["rate"],ytd_rev=t["ytd"]["revenue"],mtd_calls=t["mtd"]["calls"],
                        mtd_tgls=t["mtd"]["tgls"],mtd_rate=t["mtd"]["rate"]))
    out.sort(key=lambda x:(-x["ytd_rate"],-x["ytd_tgls"]))
    return out

def alltotals(at):
    def agg(rows):
        c=sum(x["ytd_calls"] for x in rows); g=sum(x["ytd_tgls"] for x in rows)
        mc=sum(x["mtd_calls"] for x in rows); mg=sum(x["mtd_tgls"] for x in rows)
        return dict(ytd_c=c,ytd_t=g,ytd_rate=rate(g,c),ytd_rev=sum(x["ytd_rev"] for x in rows),
                    mtd_c=mc,mtd_t=mg,mtd_rate=rate(mg,mc))
    return (agg([x for x in at if x["is_silo"]]), agg([x for x in at if not x["is_silo"]]), agg(at))

def dept_techs_list(techs):
    rows=[(n,t["ytd"]["calls"]) for n,t in techs.items() if "," not in n]
    rows.sort(key=lambda x:-x[1]); return [n for n,_ in rows]

def _row(n, t, cancel, sd, before=None):
    y=t["ytd"]; calls,tgls,rev = y["calls"],y["tgls"],y["revenue"]
    if before:
        for m in MONTH_NAMES[:before.month-1]:
            calls-=t["monthly"][m]["calls"]; tgls-=t["monthly"][m]["tgls"]; rev-=t["monthly"][m]["revenue"]
    mt=t["mtd"]; cc=cancel[n]; s=sd[n]["ytd"]; sdnd=s["flipped"]+s["nextday"]
    return dict(name=n,ytd_calls=calls,ytd_tgls=tgls,ytd_rate=rate(tgls,calls),ytd_rev=rev,
                mtd_calls=mt["calls"],mtd_tgls=mt["tgls"],mtd_rate=rate(mt["tgls"],mt["calls"]),mtd_rev=mt["revenue"],
                cancel_count=cc["ytd"],cancel_scheduled=y["tgls"],cancel_rate=rate(cc["ytd"],y["tgls"]),
                sdnd_count=sdnd,sdnd_total=s["total"],sdnd_rate=rate(sdnd,s["total"]))

def _ttot(rows):
    c=sum(r["ytd_calls"] for r in rows); g=sum(r["ytd_tgls"] for r in rows)
    mc=sum(r["mtd_calls"] for r in rows); mg=sum(r["mtd_tgls"] for r in rows); mrv=sum(r["mtd_rev"] for r in rows)
    cc=sum(r["cancel_count"] for r in rows); cs=sum(r["cancel_scheduled"] for r in rows)
    sc=sum(r["sdnd_count"] for r in rows); st=sum(r["sdnd_total"] for r in rows)
    return dict(calls=c,tgls=g,rate=rate(g,c),revenue=sum(r["ytd_rev"] for r in rows),mtd_calls=mc,mtd_tgls=mg,
                mtd_rate=rate(mg,mc),mtd_rev=mrv,cancel_count=cc,cancel_scheduled=cs,cancel_rate=rate(cc,cs),
                sdnd_count=sc,sdnd_total=st,sdnd_rate=rate(sc,st),mtd_revenue=mrv)

def build_silo_only(techs, cancel, sd):
    allr=[_row(n,techs[n],cancel,sd) for n in SILO if n in techs]
    ar=[_row(n,techs[n],cancel,sd,ROSTER_START.get(n)) for n in TEAM_A if n in techs]
    br=[_row(n,techs[n],cancel,sd,ROSTER_START.get(n)) for n in TEAM_B if n in techs]
    c=sum(r["ytd_calls"] for r in allr); g=sum(r["ytd_tgls"] for r in allr)
    mc=sum(r["mtd_calls"] for r in allr); mg=sum(r["mtd_tgls"] for r in allr)
    return {"techs":allr,"ytd_totals":dict(calls=c,tgls=g,rate=rate(g,c),revenue=sum(r["ytd_rev"] for r in allr)),
            "mtd_totals":dict(calls=mc,tgls=mg,rate=rate(mg,mc)),
            "team_a":{"techs":ar,"totals":_ttot(ar)},"team_b":{"techs":br,"totals":_ttot(br)}}

def build_yesterday(today):
    yday = today - timedelta(days=1)
    calls=defaultdict(int); tgls=defaultdict(int); rev=defaultdict(float)
    def st(name):
        out=[n for n in names_of(name) if n in SILO]
        if out: return out
        f=resolve(name); return [f] if f in SILO else []
    for grp, r in iter_grouped(load_rows("Revenue_By_JobType.xlsx"), "Assigned Technicians", 3):
        if to_date(r[4])!=yday: continue
        for tech in st(r[7]): calls[tech]+=1
    SUB = sub_by_srcjob()
    for grp, r in iter_grouped(load_rows("ROPP_TGLs_Created.xlsx"), "Assigned Technicians", 1):
        if to_date(r[5])!=yday: continue
        for tech in st(r[3]): tgls[tech]+=1; rev[tech]+=SUB.get(jobkey(r[1]), 0.0)
    def team(names):
        rows=[dict(name=n,calls=calls.get(n,0),tgls=tgls.get(n,0),rate=rate(tgls.get(n,0),calls.get(n,0)),
                   revenue=round(rev.get(n,0))) for n in names]
        rows.sort(key=lambda x:x["name"])
        tc=sum(x["calls"] for x in rows); tg=sum(x["tgls"] for x in rows)
        return {"techs":rows,"totals":dict(calls=tc,tgls=tg,rate=rate(tg,tc),revenue=sum(x["revenue"] for x in rows))}
    return {"date":yday.isoformat(),"team_a":team(TEAM_A),"team_b":team(TEAM_B)}

def build_dept_pace(allt, de):
    dr=DAYS_TOTAL-de; ytd=allt["ytd_rev"]; exp=round(DEPT_GOAL*de/DAYS_TOTAL); da=round(ytd/de) if de else 0
    return {"goal":DEPT_GOAL,"ytd_revenue":ytd,"expected_pace":exp,"ahead":ytd-exp,
            "projected_eoy":round(ytd+da*dr),"pace_pct":round(ytd/exp*100,1) if exp else 0,"daily_actual":da,
            "daily_needed":round((DEPT_GOAL-ytd)/dr) if dr else 0,"days_elapsed":de,"days_remaining":dr,
            "days_total":DAYS_TOTAL,"total_calls":allt["ytd_c"],"total_tgls":allt["ytd_t"],"conv_rate":allt["ytd_rate"],
            "rev_per_tgl":round(allt["ytd_rev"]/allt["ytd_t"]) if allt["ytd_t"] else 0,
            "silo_rev":ytd,"other_rev":0,"total_revenue":ytd}

# =============================================================================
#  INJECTION
# =============================================================================
def _end(html, i):
    oc=html[i]; cc='}' if oc=='{' else ']'; d=0; k=i; ins=esc=False
    while k<len(html):
        c=html[k]
        if ins:
            if esc:esc=False
            elif c=='\\':esc=True
            elif c=='"':ins=False
        else:
            if c=='"':ins=True
            elif c==oc:d+=1
            elif c==cc:
                d-=1
                if d==0:return k
        k+=1
    return -1

def replace_const(html, name, value):
    js=json.dumps(value,ensure_ascii=False)
    m=re.search(r'const\s+'+re.escape(name)+r'\s*=\s*',html)
    if not m: print("  (skip)",name); return html
    i=m.end()
    if html[i] in '{[': k=_end(html,i); return html[:i]+js+html[k+1:]
    j=html.index(';',i); return html[:i]+js+html[j:]

def js_obj(d): return "{"+",".join(k+":"+json.dumps(v) for k,v in d.items())+"}"
def replace_raw(html, name, raw):
    m=re.search(r'const\s+'+re.escape(name)+r'\s*=\s*',html)
    if not m: print("  (skip)",name); return html
    i=m.end()
    if html[i]=='{': k=_end(html,i); return html[:i]+raw+html[k+1:]
    j=html.index(';',i); return html[:i]+raw+html[j:]

def replace_list(html, name, lst):
    arr="["+",".join('"'+n.replace('"','\\"')+'"' for n in lst)+"]"
    m=re.search(r'const\s+'+re.escape(name)+r'\s*=\s*\[',html)
    if not m: print("  (skip)",name); return html
    i=m.end()-1; k=_end(html,i); return html[:i]+arr+html[k+1:]

def patch_sets(html):
    s15="["+",".join('"'+n+'"' for n in SILO)+"]"
    for nm in ("SILO_SET","SILO_SET_R"):
        html=re.sub(r'(const\s+'+nm+r'\s*=\s*new Set\()\[[^\]]*\](\s*\))', lambda m:m.group(1)+s15+m.group(2), html)
    html=re.sub(r'(const\s+techList\s*=\s*)\[[^\]]*\]', lambda m:m.group(1)+s15, html)
    return html

def find_template(explicit):
    if explicit and (SCRIPT_DIR/explicit).exists(): return SCRIPT_DIR/explicit
    cands=sorted(SCRIPT_DIR.glob("ROPP_Dashboard_2026_DEPARTMENT_Sierra_*.html"),key=lambda p:p.stat().st_mtime,reverse=True)
    if not cands:
        cands=sorted(SCRIPT_DIR.glob("ROPP_Dashboard_2026_*Sierra_*.html"),key=lambda p:p.stat().st_mtime,reverse=True)
    if not cands: raise SystemExit("No dashboard template found.")
    return cands[0]

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--date"); ap.add_argument("--template"); ap.add_argument("--out")
    a=ap.parse_args()
    today=datetime.strptime(a.date,"%Y-%m-%d").date() if a.date else date.today()
    de=(today-YEAR_START).days; months=MONTH_NAMES[:today.month]
    print("Rebuilding DEPARTMENT dashboard for", today, "(day", de, ")")
    techs, cancel, sd = parse_all(today)
    names=[n for n in techs if "," not in n]
    AT = build_allteams(techs)
    at_s, at_n, at_a = alltotals(AT)
    blocks = {
        "INITIAL_DATA": build_initial(techs),
        "CANCEL_DATA": build_cancel(techs,cancel,today,months),
        "SAMEDAY_DATA": build_sameday(sd,today,months,names),
        "WEEKLY_DATA": build_weekly_prevmonth(today),
        "WEEKLY_CONV_DATA": build_weekly_conv(today,names),
        "PACE_DATA": build_pace(techs,today,de),
        "MONTHLY_DETAIL": build_monthly_detail(techs,months),
        "ALLTEAMS_DATA": AT,
        "SILO_ONLY_DATA": build_silo_only(techs,cancel,sd),
        "YESTERDAY_DATA": build_yesterday(today),
        "DEPT_PACE_DATA": build_dept_pace(at_a,de),
    }
    tmpl=find_template(a.template); print("Template:",tmpl.name)
    html=tmpl.read_text(encoding="utf-8")
    for nm,val in blocks.items(): html=replace_const(html,nm,val)
    html=replace_raw(html,"ALLTEAMS_SILO_TOT",js_obj(at_s))
    html=replace_raw(html,"ALLTEAMS_NS_TOT",js_obj(at_n))
    html=replace_raw(html,"ALLTEAMS_ALL_TOT",js_obj(at_a))
    html=replace_list(html,"SILO_12",SILO_12)
    html=replace_list(html,"DEPT_TECHS",dept_techs_list(techs))
    html=patch_sets(html)
    cur=today.strftime("%b ")+str(today.day)+", "+str(YEAR)
    html=re.sub(r'Jan 1 [-–] [A-Z][a-z]{2} \d{1,2}, '+str(YEAR), "Jan 1 – "+cur, html)
    out=a.out or ("ROPP_Dashboard_2026_DEPARTMENT_Sierra_"+today.strftime('%b%d')+".html")
    (SCRIPT_DIR/out).write_text(html,encoding="utf-8")
    print("DEPT YTD:", at_a['ytd_c'], "calls /", at_a['ytd_t'], "TGLs /", at_a['ytd_rate'], "%  MTD",
          at_a['mtd_c'], "/", at_a['mtd_t'], "/", at_a['mtd_rate'], "%")
    print("Saved:", SCRIPT_DIR/out)

if __name__ == "__main__":
    main()
